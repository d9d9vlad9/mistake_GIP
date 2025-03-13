import openpyxl
from openpyxl import Workbook

def log_message_to_excel(sheet_name, data, save_interval=1000):
    """
    Записывает сообщения на указанный лист в файле Excel.

    :param sheet_name: Имя листа, на который нужно записать сообщение.
    :param data: Список кортежей, где первый элемент — это имя файла (filename), а второй — сообщение (message).
                 Для листа 'Созданные файлы' filename может быть None.
    :param save_interval: Количество записей, после которого производится сохранение файла.
    """
    try:
        try:
            workbook = openpyxl.load_workbook('log_results.xlsx')
        except FileNotFoundError:
            workbook = Workbook()
            if 'Sheet' in workbook.sheetnames:
                std_sheet = workbook['Sheet']
                workbook.remove(std_sheet)

        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)

        sheet = workbook[sheet_name]


        if sheet.max_row == 0 or (sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None):
            if sheet_name == 'Созданные файлы':
                sheet.append(['Filename'])
            elif sheet_name == 'Подписанные файлы':
                sheet.append(['Filename'])
            else:
                sheet.append(['Local_uid', 'Message'])

        for i, entry in enumerate(data, start=1):
            filename, message = entry
            if sheet_name == 'Созданные файлы':
                sheet.append([filename])
            elif sheet_name == 'Подписанные файлы':
                sheet.append([filename])  
            else:
                sheet.append([filename, message])

            if i % save_interval == 0:
                workbook.save('log_results.xlsx')

        workbook.save('log_results.xlsx')

    except Exception as e:
        print(f"Error occurred: {e}. Data might be partially saved.")
